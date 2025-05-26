#Interfaz gráfica
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
import customtkinter as ctk
from PIL import Image, ImageTk 

import os
import json
#hilos para evitar el crasheo  de la ui
import threading

user_pc = os.environ["USERPROFILE"]
user_name = os.path.basename(user_pc) 
name = user_name.split(".")
name = str(name[0]).capitalize()

base_path = fr"{user_pc}\35159_147728_DUPREE_VENTA_DIRECTA_S_A\MERCADEO PAISES - MERCADEO PAISES DOCUMENTOS\Mercadeo BI\1. Arquitectura\Bases\15. Base Revisión LLL" 
secret_path = fr"{user_pc}\35159_147728_DUPREE_VENTA_DIRECTA_S_A\MERCADEO PAISES - MERCADEO PAISES DOCUMENTOS\Mercadeo BI\1. Arquitectura\Bases\1. Plantillas\1. Regional\RegistroSTau"
params_path = fr"{base_path}\2. Imágenes\params\params.json"
version_path = fr"{base_path}\2. Imágenes\version.json"

with open(params_path, "r", encoding="utf-8") as f:
    params = json.load(f)

with open(version_path, "r", encoding="utf-8") as f:
    version = json.load(f)

version_actual = "1.0"
version_sp = version.get("version")

files_loaded = False

version_path = fr"{base_path}/2. Imágenes/version.json"

if os.path.exists(version_path):
    with open(version_path, "r") as f:
        version_data = json.load(f)
        file_version = version_data.get("version", None)

#Variable global para insertar texto en la consola
consola = None

#Archivo seleccionado
selected_file = None

#Variables de estilos
font_h1 = "Gilroy Black", 16
font_h2 = "Gilroy Black", 12 
font_normal = "Gilroy Medium", 14
font_consola = "Consolas", 12, "bold"

background_color = "#F7FAFA"
text_color = "#202C39"
#secondary_button = "#228CDB"
secondary_button = "#957FEF"
disabled_color = "#7C8483"

def main():
    ui()

#actualizar mensaje consola
def actualizar_consola(mensaje):
    consola.configure(state="normal")         
    consola.insert("end", f">>{mensaje}\n\n")     
    consola.see("end")
    consola.configure(state="disabled")
    ventana.update_idletasks()

def cargar_xlsx():
    global consola, btn_evaluar_lll, selected_file, files_loaded

    # Limpiar consola
    consola.configure(state="normal")
    consola.delete("1.0", "end")
    bienvenida = f">>¡Hola {name}! Te doy la \nbienvenida a Gravitón.\nPor favor, comienza cargando\nel LLL a revisar...\n\n"
    consola.insert("0.0", bienvenida)
    consola.configure(state="disabled")

    print("Cargando Leader List Lite...")

    # Abre explorador de archivos para un solo archivo
    path = filedialog.askopenfilename(
        title="Seleccione el archivo LLL",
        filetypes=[("Archivos Excel", ("*.xlsb", "*.xlsx"))]
    )

    if path:
        selected_file = path
        files_loaded = True

        archivo_msg = f"{name}, el archivo elegido\nfue: {os.path.basename(path)}"
        actualizar_consola(archivo_msg)

        activar_analisis()
    else:
        print("No se cargó ningún archivo.")

#detector de errores nativos de excel
def es_error_excel(v):
    return isinstance(v, str) and v.startswith("#") and v.endswith("!")

def configurar_logs(user_name, base_path, secret_path):
    import logging
    logger = logging.getLogger("graviton_logger")
    logger.setLevel(logging.INFO)

    if not logger.handlers:
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

        # Ruta 1: log público
        ruta_log_publico = os.path.join(base_path, "3. Log de Registro", f"log_graviton_{user_name}.log")
        os.makedirs(os.path.dirname(ruta_log_publico), exist_ok=True)
        handler_publico = logging.FileHandler(ruta_log_publico, encoding="utf-8")
        handler_publico.setFormatter(formatter)
        logger.addHandler(handler_publico)

        # Ruta 2: log secreto
        ruta_log_secreto = os.path.join(secret_path, "Log", f"log_graviton_{user_name}.log")
        os.makedirs(os.path.dirname(ruta_log_secreto), exist_ok=True)
        handler_secreto = logging.FileHandler(ruta_log_secreto, encoding="utf-8")
        handler_secreto.setFormatter(formatter)
        logger.addHandler(handler_secreto)

    return logger

def ui():
    import datetime
    global consola, btn_evaluar_lll, files_loaded, pais, campana, ventana, btn_descargar, btn_cargar_lll, ruta_ico

    ventana = tk.Tk()
    ventana.title("Gravitón - Verificador de LLL")
    ventana.geometry("550x450")
    ventana.configure(bg=background_color)
    ventana.attributes('-alpha', 0.98)
    ruta_ico = fr"{base_path}/2. Imágenes/favicon.ico"

    #Estilos
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")          

    try:
        ventana.iconbitmap(ruta_ico) 
    except Exception as e:
        messagebox.showerror("Error de sincronización ♾️", "Marketing Países / Mercadeo BI se encuentra sincronizada de manera erronea.")
    
    ventana.resizable(False, False)    

    titulo = tk.Label(ventana, text="Gravitón — Verificador de Leader List Lite", font=font_h1, bg=background_color, fg=text_color)
    titulo.pack(pady=(15,0))

    #Establecer dos columnas
    frame_contenido = tk.Frame(ventana, bg=background_color)
    frame_contenido.pack(fill="both", expand=True, padx=20, pady=5)

    #----------------------Columna izquierda
    columna_izquierda = tk.Frame(frame_contenido, bg=background_color)
    columna_izquierda.grid(row=0, column=0, sticky="nsew")

    # Columna derecha
    columna_derecha = tk.Frame(frame_contenido, bg=background_color)
    columna_derecha.grid(row=0, column=1, sticky="nsew")

    # Configurar proporción de columnas
    frame_contenido.columnconfigure(0, weight=1)
    frame_contenido.columnconfigure(1, weight=1)    

    #both 
    frame_both = tk.Frame(ventana, bg=background_color)
    frame_both.pack(fill="both", expand=True, padx=5, pady=5)  

    #Contenido en izquierda
    #--------------------------DropDowns
    dropdown_frame = ctk.CTkFrame(
        columna_izquierda,
        fg_color="transparent",
        corner_radius=0,
        width=200            
    )
    # evitar cambio de tamaño
    dropdown_frame.pack(pady=(16, 0), anchor="center")
    dropdown_frame.pack_propagate(False)

    # grid
    dropdown_frame.grid_columnconfigure(0, weight=3)   # 30% del espacio
    dropdown_frame.grid_columnconfigure(1, weight=7)   # 70% del espacio
    dropdown_frame.grid_rowconfigure(0, weight=1)    

    #pais
    pais = tk.StringVar(value="País")    
    pais.trace_add("write", lambda *args: activar_analisis())
    paises = params.get("paises")

    pais_dropdown = ctk.CTkOptionMenu(
        dropdown_frame,
        variable=pais,
        values=paises,
        fg_color=text_color,
        button_color=text_color,
        text_color=background_color,
        dropdown_fg_color=text_color,
        dropdown_text_color=background_color,
        corner_radius=6,
        height=30,
        width=80,
        dynamic_resizing=False,
        font=font_normal
    )
    pais_dropdown.grid(row=0, column=0, sticky="ew", padx=(0,5))    

    #campaña
    campana = tk.StringVar(value="Campaña")
    campana.trace_add("write", lambda *args: activar_analisis())
    campana_inicial = params.get("camp_inicial")
    fecha = datetime.date.today()
    mes = fecha.month
    anio = fecha.year
    #mes = 12

    camp  = mes + 7
    if camp > 18:
        camp -= 18
        anio += 1
    campana_final   = anio * 100 + camp

    #indices
    anio_inicial = campana_inicial // 100
    periodo_inicial = campana_inicial % 100
    idx_inicial = anio_inicial*18 + (periodo_inicial - 1)
    anio_final = campana_final // 100
    periodo_final = campana_final % 100
    idx_final = anio_final*18 + (periodo_final - 1)

    #range
    idxs = range(idx_inicial, idx_final + 1)
    campanas = [
        (i // 18)*100 + (i % 18 + 1)
        for i in idxs
    ]
    campanas.sort(reverse=True)  # para dropdown de mayor a menor

    campana_dropdown = ctk.CTkOptionMenu(
        dropdown_frame,
        variable=campana,
        values=[str(c) for c in campanas],
        fg_color=text_color,
        button_color=text_color,
        text_color=background_color,
        dropdown_fg_color=text_color,
        dropdown_text_color=background_color,
        corner_radius=6,
        height=30,
        width=110,
        dynamic_resizing=False,
        font=font_normal
    )
    campana_dropdown.grid(row=0, column=1, sticky="ew", padx=(5,0))

    #botones
    btn_cargar_lll = ctk.CTkButton(
        columna_izquierda, 
        text="Cargar Leader List Lite",
        command=cargar_xlsx,
        fg_color=secondary_button,
        text_color=text_color,
        font=font_normal,
        corner_radius=12,
        height=60, 
        width=200        
    )
    btn_cargar_lll.pack(padx=0, pady=(20, 0), anchor="center")    

    #botones
    btn_evaluar_lll = ctk.CTkButton(
        columna_izquierda, 
        text="Comenzar evaluación de\nLLL cargados",
        command=thread_evaluacion,
        #fg_color=items_color,
        fg_color = disabled_color,
        text_color=background_color,
        font=font_normal,
        corner_radius=12,
        height=60, 
        width=200,
        state="disabled"        
    )
    btn_evaluar_lll.pack(padx=0, pady=(20,20), anchor="center")

    btn_descargar = ctk.CTkButton(
        columna_izquierda, 
        text="Descargar resultados",
        command=generar_archivo,
        #fg_color="#005E9E",
        fg_color= disabled_color,
        text_color = background_color,
        font=font_normal,
        corner_radius=12,
        height=60, 
        width=200,
        state="disabled"        
    )
    btn_descargar.pack(padx=0, pady=0, anchor="center")

    #----------------------Columna izquierda
    
    consola = ctk.CTkTextbox(
    columna_derecha,
    corner_radius=12,
    fg_color=text_color,
    text_color=background_color,
    #text_color=background_color,
    font=font_consola,
    height=345
    )
    consola.pack(pady=(16,0), padx=(0,20), fill="both", expand=True)  

    bienvenida = f">>¡Hola {name}! Te doy la \nbienvenida a Gravitón.\nPor favor, comienza cargando\nel LLL a revisar...\n\n"
    consola.insert("0.0", bienvenida)

    consola.configure(state="disabled")

    #*----------------------Logo IN sobrepuesta
    # Carga y redimensiona la imagen a 10x10 píxeles
    try:
        imagen_original = Image.open(f"{base_path}/2. Imágenes/logo.png")
    except Exception as e:
        messagebox.showwarning("Error", "Los iconos no pueden ser encontrados, por favor, sincronice sus carpetas.")

    imagen_redimensionada = imagen_original.resize((130, 62), resample=Image.Resampling.LANCZOS)
    imagen_tk = ImageTk.PhotoImage(imagen_redimensionada)

    # Crea el Label con la imagen
    lbl_imagen = tk.Label(ventana, image=imagen_tk, bd=0, bg=background_color)
    lbl_imagen.image = imagen_tk 

    # Delante de todo y coordenadas
    lbl_imagen.place(relx=1.0, y=350, x=-348, anchor="ne")    
    lbl_imagen.lift()

    ventana.mainloop()

def reiniciar_gui():
    global files_loaded, selected_file, errores_lll
    # Reiniciar variables globales
    files_loaded = False
    selected_file = None
    errores_lll = []
    
    # Restablecer menús desplegables
    pais.set("País")
    campana.set("Campaña")
    
    # Limpiar consola
    consola.configure(state="normal")
    consola.delete("1.0", "end")
    bienvenida = f">>¡Hola {name}! Te doy la \nbienvenida a Gravitón.\nPor favor, comienza cargando\nel LLL a revisar...\n\n"
    consola.insert("0.0", bienvenida)
    consola.configure(state="disabled")
    
    # Desactivar botones
    btn_evaluar_lll.configure(state="disabled", fg_color=disabled_color, text_color=text_color)
    btn_descargar.configure(state="disabled", fg_color=disabled_color, text_color="#E6E6E6")

def popup(titulo, texto):
    dlg = ctk.CTkToplevel(ventana)
    dlg.title(titulo)
    ancho, alto = 300, 150
    dlg.geometry(f"{ancho}x{alto}")

    # 2) No dejar cambiar tamaño
    dlg.resizable(False, False)
    # (opcional) quitar maximizar/minimizar
    dlg.attributes('-toolwindow', True)

    dlg.attributes('-topmost', True)
    dlg.grab_set()    

    # Label con wrap y centrado dentro del fijo
    lbl = ctk.CTkLabel(
        dlg,
        text=texto,
        font=font_normal,
        wraplength=ancho - 40,  # deja algo de margen
        justify="center"
    )
    lbl.place(relx=0.5, rely=0.4, anchor="center")

    btn = ctk.CTkButton(
        dlg,
        text="OK",
        command=dlg.destroy,
        fg_color=secondary_button,
        text_color=text_color,
        font=font_normal,
        width=120,
        height=30
    )
    btn.place(relx=0.5, rely=0.75, anchor="center")

    # Centrar ventana sobre la principal
    px = ventana.winfo_rootx()
    py = ventana.winfo_rooty()
    pw = ventana.winfo_width()
    ph = ventana.winfo_height()
    x = px + (pw - ancho)//2
    y = py + (ph - alto)//2
    dlg.geometry(f"{ancho}x{alto}+{x}+{y}")

    dlg.wait_window()

def thread_evaluacion():
    #evitar doble clic
    btn_cargar_lll.configure(state="disabled")
    btn_evaluar_lll.configure(state="disabled")
    threading.Thread(target=_verificar_con_callback, daemon=True).start()

def thred_generacion():
    btn_descargar.configure(state="disabled")
    threading.Thread(target=_generar_con_callback, daemon=True).start()

def _generar_con_callback():
    try:
        generar_archivo()
    finally:
        ventana.after(0, lambda: btn_descargar.configure(state="normal"))

def _verificar_con_callback():
    try:
        verificar_seleccion()
    finally:
        #cuando termine, vuelve al hilo de la GUI para reactivar
        ventana.after(0, lambda: btn_evaluar_lll.configure(state="normal"))
        ventana.after(0, lambda: btn_cargar_lll.configure(state="normal"))

def leer_lll(selected_file, sheets):
    import pandas as pd
    ext = os.path.splitext(selected_file)[-1].lower()
    if ext == ".xlsb":
        df = pd.read_excel(selected_file, sheet_name=sheets, header=None, engine="pyxlsb")
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(selected_file, sheet_name=sheets, header=None, engine="openpyxl")  
    else:
        raise ValueError("Formato de archivo no soportado") 

    return df

def activar_analisis():
    global btn_evaluar_lll, files_loaded, pais_seleccionado, campana_seleccionada
    #get valores
    pais_seleccionado = pais.get()
    campana_seleccionada = campana.get()   

    #condiciones
    if files_loaded and pais_seleccionado != "País" and campana_seleccionada != "Campaña":
        btn_evaluar_lll.configure(state="normal", fg_color=text_color, text_color=background_color)
    else:
        btn_evaluar_lll.configure(state="disabled", fg_color=disabled_color)

def acivar_resultado():
    if len(errores_lll) > 0:
        btn_descargar.configure(state="normal", fg_color=secondary_button, text_color=text_color)

def verificar_seleccion():
    global campana_B5, pais_L5

    try:
        df_CAT = leer_lll(selected_file, "LL CAT")    
    except Exception as e:
        messagebox.showerror("Error en la lectura", f"El archivo cargado no tiene la estructura adecuada.\n{name}, asegurate de usar la versión más reciente del formato de LLL entregado por el Área de Precios y Optimización y de no modificar los nombres y/o formatos de las primeras 71 columnas.")
    pais_L5 = df_CAT.iloc[4,11]
    campana_B5 = df_CAT.iloc[4,1]  

    if str(pais_seleccionado) == str(pais_L5) and str(campana_seleccionada) == str(campana_B5):
        validar_lll()
    else:
        messagebox.showerror("Error en selección", "El país o campaña seleccionados no coinciden con el país o campaña del LLL cargado.")
        reiniciar_gui()

def validar_lll():
    import pandas as pd
    global selected_file, ventana, errores_lll, df_CAT, columnas_lll, campana_B5, pais_L5, nombre_archivo
    
    #print("País seleccionado: ", pais_seleccionado)
    #print("Campaña: ", campana_seleccionada)

    print("Comenzando validación")        

    #almacen de errores
    errores_lll = []
    fila_inicial = 6

    #encontrar el final del LLL
    def encontrar_fin_datos(df, fila_inicial):
        #buscar la primera fila completamente vacía desde fila_inicial
        for idx in range(fila_inicial, len(df)):
            if df.iloc[idx].isna().all():
                return idx  #devolver el indice de la ultima fila
        return len(df) #si no lo encuentra devolver el len del df    
        
    #validar igual a cero
    def validar_cero(columna:int):
        for idx in range(fila_inicial, fila_final):
            valor = df_CAT.iat[idx, columna]
            if valor == 0:
                errores_lll.append({'fila': idx, 'col': columna})

    #validar vacios o errores
    def validar_vacios(columna:int):
        for idx in range(fila_inicial, fila_final):
            valor = df_CAT.iat[idx, columna]
            # Verificar si el valor es NaN
            if pd.isna(valor):
                errores_lll.append({'fila': idx, 'col': columna})
                continue
            
            # Si es una cadena, verificar si está vacía o contiene un error
            if isinstance(valor, str):
                valor_stripped = valor.strip()
                if valor_stripped == "" or valor_stripped.startswith("#"):
                    errores_lll.append({'fila': idx, 'col': columna})

    #valores repetidos en la misma columna
    def validar_duplicados(columna:int):
        valores_vistos = set() #almacen

        for idx in range(fila_inicial, fila_final):
            valor = df_CAT.iat[idx, columna]

            # Verificar duplicados
            if valor in valores_vistos:
                errores_lll.append({'fila': idx, 'col': columna})
            else:
                valores_vistos.add(valor)

    #validar columnas con arrays de params, validar que este in, que no sean vacío u error
    def validacion_basica(columna:int, array_evaluativo):
        
        for idx in range(fila_inicial, fila_final):
            valor = df_CAT.iat[idx, columna]
            #celda vacía o NaN
            if pd.isna(valor) or (isinstance(valor, str) and valor.strip() == ""):
                errores_lll.append({'fila': idx, 'col': columna})

            #error de Excel (#REF!, #N/A…)
            if es_error_excel(valor):
                errores_lll.append({'fila': idx, 'col': columna})
                continue

            #validación contra validos
            v_str = str(valor).strip()
            if v_str not in array_evaluativo:
                errores_lll.append({'fila': idx, 'col': columna})

    def validar_rango_paginas(pagina_evaluada: int):
        col_vehiculo = 1 
        for idx in range(fila_inicial, fila_final):
            vehiculo = df_CAT.iat[idx, col_vehiculo]
            pagina = df_CAT.iat[idx, pagina_evaluada]

            try:
                #convertir int
                num_pagina = int(pagina)

                #filtrar rango del vehículo
                rango_vehiculo = rango_paginas[rango_paginas['codi_vehi'] == vehiculo]

                if rango_vehiculo.empty:
                    errores_lll.append({'fila': idx, 'col': col_vehiculo})
                else:
                    #obtener límites
                    pag_inicial = rango_vehiculo['PAG INICIAL'].values[0]
                    pag_final = rango_vehiculo['PAG FINAL'].values[0]

                    #validar rango
                    if not (pag_inicial <= num_pagina <= pag_final):
                        errores_lll.append({'fila': idx, 'col': pagina_evaluada})

            except Exception as e:
                errores_lll.append({'fila': idx, 'col': pagina_evaluada})

    def consumir_conSQL(columna_inicial, columna_final, filas:int):
        #toca asi porque no es capaz con la hoja completa
        df_conSQL = pd.read_excel(selected_file,
            sheet_name='Consultas SQL',
            engine='pyxlsb',
            header=1, #primera fila es header
            usecols=f"{columna_inicial}:{columna_final}", nrows=int(filas)) #indicativo columnas
        
        df_conSQL = df_conSQL.dropna(how='all')
        #print(df_conSQL)
        return df_conSQL
    
    actualizar_consola("Comenzando el análisis del\nLLL cargado...")   

    if not selected_file:
        messagebox.showerror("Sin archivo", "No se ha seleccionado un archivo, por favor, seleccione alguno antes de continuar")
        return
                
    #carga  
    try:
        df_CAT = leer_lll(selected_file, "LL CAT")
    except Exception as e:
        messagebox.showerror("Error al leer archivo", f"El archivo cargado no corresponde a un LLL")

        return

    #-----------------Validar que sea un LLL a través de las primeras 71 columnas
    columnas_lll = params.get("cols_LLL")
    fila_final = encontrar_fin_datos(df_CAT, fila_inicial)

    #print(columnas_lll)
    
    #Definición headers
    #print("Archivo cargado")
    headers = df_CAT.iloc[5].fillna("").tolist()
    uens_detectadas = df_CAT.iloc[fila_inicial:fila_final, 4].unique()

    #Si coinciden los headers significa que es un LLL
    if headers[:71] != columnas_lll:
        messagebox.showerror("Error en la lectura", "El LLL seleccionado no tiene la estructura adecuada. Por favor, no modificar el formato original entregado por el Área de Precios y Optimización. Específicamente por Don Rodrigo")
        """consola.configure(state="normal")         
        consola.insert("end", "\n>>ERROR: El archivo cargado no corresponde a un LLL...\n\n")     
        consola.see("end")
        consola.configure(state="disabled") """

    #----------------Validar que la campaña y el país sea igual
    if str(campana_B5) != str(campana_seleccionada):
        messagebox.showerror("Disparidad en campañas", "La campaña seleccionada no corresponde a la campaña del LLL cargado.")        

    if str(pais_L5) != str(pais_seleccionado):
        messagebox.showerror("Disparidad en país", "El país seleccionado no corresponde al país del LLL cargado.")        
    
    actualizar_consola("Validando filas...")

    #--------------------Validar Tipo de Venta -- CAMBIAR

    uens_especiales = params.get("uens_especiales")
    print(uens_especiales)
    print(uens_detectadas)
    if any(uen in uens_especiales for uen in uens_detectadas):
        tipos_venta = consumir_conSQL("Y", "Y", 40)
        print("Sí hay UENs especiales")
    else:
        tipos_venta = consumir_conSQL("X", "X", 40)
        print("No hay UENs especiales")

    tipos_venta = tipos_venta['Tipo Venta'].tolist()
    tipos_venta = [int(x) for x in tipos_venta]
    tipos_venta = [str(x) for x in tipos_venta]
    print(tipos_venta)
    validacion_basica(0, tipos_venta)

    #--------------------Validar Vehiculo
    vehiculo = consumir_conSQL("C", "C", 50)
    vehiculo = vehiculo["codi_vehi"].tolist()
    validacion_basica(1, vehiculo)

    #--------------------Validar Estrategía
    codi_estrategia = consumir_conSQL("H", "H", 200)
    codi_estrategia = codi_estrategia["codi_estr"].tolist()
    validacion_basica(2, codi_estrategia)

    #--------------------Validar LINEA DE VENT
    validar_vacios(3)

    validar_vacios(4)

    validar_vacios(5)

    validar_vacios(6)

    validar_vacios(7)

    validar_vacios(8)

    validar_vacios(9)

    validar_vacios(10)

    validar_vacios(11)
    
    #--------------------Validar Tipo Programación
    if any(uen in uens_especiales for uen in uens_detectadas):
        tipo_prog = consumir_conSQL("BG", "BG", 15)
        print("Sí hay UENs especiales")
    else:
        tipo_prog = consumir_conSQL("BF", "BF", 15)
        print("No hay UENs especiales")
        
    tipo_prog = tipo_prog["Tipo_Prog"].tolist()
    validacion_basica(12, tipo_prog)

    #--------------------Validar COLECCIÓN CAPSULA
    validar_vacios(13)

    #--------------------Validar PAG NAL
    rango_paginas = consumir_conSQL("A", "F", 40)
    #print(rango_paginas)
    validar_vacios(14)
    validar_rango_paginas(14)

    #--------------------Validar PAG ADV
    validar_vacios(15)
    validar_rango_paginas(15)

    #--------------------Validar COD VENTA
    validar_vacios(18)
    validar_duplicados(18)
    
    #--------------------Validar COD PROD
    validar_vacios(19)
    validar_duplicados(19)

    #--------------------Validar DESCRIPCIÓN COMERCIAL PAÍS
    validar_vacios(20)
    #validar_duplicados(20)

    validar_vacios(21)

    validar_vacios(22)

    validar_vacios(23)

    #--------------------Validar OBSERVACION PDF
    if any(uen in uens_especiales for uen in uens_detectadas):
        valores_pdf = consumir_conSQL("CK", "CK", 15)
        print("Sí hay UENs especiales")
    else:
        valores_pdf = consumir_conSQL("CJ", "CJ", 15)
        print("No hay UENs especiales")

    valores_pdf = valores_pdf.iloc[:, 0].tolist()
    validacion_basica(26, valores_pdf)

    #--------------------Validar COSTO INICIAL
    validar_vacios(29)

    #--------------------Validar CANASTA FINAL
    validar_vacios(30)

    validar_vacios(32)

    #--------------------Validar FLETE
    validar_vacios(33)

    #--------------------Validar PP
    validar_vacios(34)

    #--------------------Validar PO
    validar_vacios(35)

    #--------------------Validar DCTO ASESORA
    validar_vacios(37)

    validar_vacios(38)

    validar_vacios(39)

    #--------------------Validar V NETO
    validar_vacios(40)

    #--------------------Validar FACTOR INICIAL/BASE
    validar_vacios(41)

    #--------------------Validar AJUSTE FACTOR
    for idx in range(fila_inicial, fila_final):
        valor = float(df_CAT.iat[idx, 42])
        if valor < 0 or valor > 2:
            errores_lll.append({'fila': idx, 'col': 42})

    #--------------------Validar FACTOR
    validar_vacios(43)
    validar_cero(43)

    for idx in range(fila_inicial, fila_final):
        valor_fact_ini = float(df_CAT.iat[idx, 41])
        valor_ajuste = float(df_CAT.iat[idx, 42])
        valor_factor = float(df_CAT.iat[idx, 43])

        factor_esperado = valor_fact_ini * valor_ajuste

        if valor_factor != factor_esperado:
            errores_lll.append({'fila': idx, 'col': 43})
            

    #--------------------Validar UNIDAD NAL
    validar_vacios(44)

    #--------------------Validar UNIDAD ADV
    validar_vacios(45)

    #--------------------Validar UNIDADES TOTALES
    validar_vacios(46)

    #--------------------Validar INV SOBRANTE A C+3 --Pendiente no captura el ref
    validar_vacios(47)
    validar_cero(47)

    validar_vacios(48)
    validar_cero(48)

    validar_vacios(49)
    validar_vacios(50)
    #--------------------Validar % CMV
    validar_vacios(51)
    #validar el procentaje
    for idx in range(fila_inicial, fila_final):
        valor = df_CAT.iat[idx, 51]    
        try:
            # Convertir diferentes formatos a float
            if isinstance(valor, str):
                # Quitar caracteres no numéricos excepto . y %
                valor_limpio = valor.strip().replace(',', '.').replace('%', '')
                if not valor_limpio:
                    raise ValueError("Vacío")

                porcentaje = float(valor_limpio)

                # Verificar si tenía % para ajustar escala
                if '%' in valor:
                    porcentaje = porcentaje  # 50% = 50.0
                else:
                    porcentaje = porcentaje * 100  # 0.5 se considera 50%
            else:
                porcentaje = float(valor) * 100  # Si es decimal (0.37 -> 37%)

            # Validar rango
            if porcentaje <= 0 or porcentaje > 100:
                errores_lll.append({'fila': idx, 'col': 51})

        except (ValueError, TypeError, AttributeError):
            # Registrar errores de conversión
            print("Fallo en la evaluación del CMV")
    
    actualizar_consola("Validando columnas...")
      
    #--------------------Validar COD ESTIMACIÓN
    #función para calcular valor esperado según la lógica de excel
    def calcular_valor_estimacion(row):
        try:
            if (
                (row[2] == "LIQUID") or       
                (row[1] == "WEB") or          
                (row[2] == "SORPRE") or       
                (row[58] == "No")             
            ):
                return 63
            elif row[2] == "SUST":            
                return 50
            elif row[1] == "PREM":           
                return 36
            else:
                return 42
            
        except Exception as e:
            print(f"Error en cálculo: {str(e)}")
            return None

    for idx in range(fila_inicial, fila_final):
        try:
            fila = df_CAT.iloc[idx] 
            valor_real = fila.iloc[53]  
            valor_esperado = calcular_valor_estimacion(fila)

            if valor_esperado is not None and valor_real != valor_esperado:
                errores_lll.append({'fila': idx, 'col': 53})  

        except:
            errores_lll.append({'fila': idx, 'col': 53})  

    #--------------------Validar AGOTAR EXISTENCIA
    def calcular_valor_agtext(row):
        try:
            if row.iloc[2] in ["WEB", "OUTLET", "ALTERN", "LIQUID"]:
                return "S"
            else:
                return "N"
        except:
            return None

    for idx in range(fila_inicial, fila_final):
        fila = df_CAT.iloc[idx]
        valor_real = fila.iloc[54]  # Índice de la columna a validar
        valor_esperado = calcular_valor_agtext(fila)
        if valor_esperado is not None and valor_real != valor_esperado:
            errores_lll.append({'fila': idx, 'col': 54})

    #--------------------Validar DIGITABLE
    def calcular_valor_digitable(row):
        try: 
            col_C = 2   
            col_S = 18   
            col_T = 19   
            col_AA = 26  
            
            # Condiciones principales
            condicion_principal = (
                row.iloc[col_C] in ["SUST", "GRATIS", "GRAMON", "1X2X", "ARMAOF", "1X2X3X"] or
                (row.iloc[col_C] in ["ASPACK", "PACK"] and row.iloc[col_S] != row.iloc[col_T])
            )

            # Lógica completa
            if condicion_principal:
                return "N"
            else:
                return "B" if row.iloc[col_AA] == "Eliminar" else "S"

        except Exception as e:
            print(f"Error en cálculo DIGITABLE: {str(e)}")
            return None

    for idx in range(fila_inicial, fila_final):
        fila = df_CAT.iloc[idx]  # Obtener fila completa
        valor_real = fila.iloc[55]  # Columna 55 (índice 54 en base 0)
        valor_esperado = calcular_valor_digitable(fila)
        if valor_esperado is not None and valor_real != valor_esperado:
            errores_lll.append({'fila': idx, 'col': 55}) 

    #--------------------Validar ORIGEN
    validar_vacios(56)

    #--------------------Validar PEDIDOS 
    if any(uen in uens_especiales for uen in uens_detectadas):
        pedidos = consumir_conSQL("AL", "AO", 30)
        print("Sí hay UENs especiales")
    else:
        pedidos = consumir_conSQL("AK", "AN", 30)
        print("No hay UENs especiales")

    pedidos['codi_camp'] = pedidos["codi_camp"].astype(int)
    pedidos['tota_pedi'] = pedidos["tota_pedi"].astype(int)
    pedidos = pedidos[pedidos['codi_camp'] == int(campana_seleccionada)]['tota_pedi'].values[0]

    pedidos = pedidos

    for idx in range(fila_inicial, fila_final):
        valor = float(df_CAT.iat[idx, 57])
        if valor != pedidos:
            errores_lll.append({'fila': idx, 'col': 57})

    #--------------------Validar INDICADOR
    indicador = set(params.get("indicador"))
    validacion_basica(58, indicador)    

    uens_aprobadas = params.get("uens_aprobadas")
    print(uens_detectadas)
    print(uens_aprobadas)

    if any(uen in uens_aprobadas for uen in uens_detectadas):
        print("¡Se encontró al menos una UEN aprobada!")
        #--------------------Validar GRAMAJE
        validar_vacios(59)
        #--------------------Validar UNIDAD DE MEDIDA
        if any(uen in uens_especiales for uen in uens_detectadas):
            unidades_medida = consumir_conSQL("CI", "CI", 50)
            print("Sí hay UENs especiales")
        else:
            unidades_medida = consumir_conSQL("CH", "CH", 50)
            print("No hay UENs especiales")

        unidades_medida = unidades_medida.iloc[:, 0].tolist()
        print(unidades_medida)
        validacion_basica(60, unidades_medida)
        #--------------------Validar PUM
        validar_vacios(61)

    validar_vacios(66)
    validar_vacios(67)
    #--------------------Validar PUNTOS
    puntos = set(params.get("puntos"))
    validacion_basica(68, puntos)    

    #--------------------Validar CAMPAÑA
    for idx in range(fila_inicial, fila_final):
        valor = float(df_CAT.iat[idx, 69])
        if valor != int(campana_seleccionada):
            errores_lll.append({'fila': idx, 'col': 69})

    #--------------------Validar PUNTOS/ MAXIPUNTAJE
    for idx in range(fila_inicial, fila_final):
        valor = df_CAT.iat[idx, 70]
        # Verificar si es entero (int) o float equivalente a entero
        es_entero = (
            isinstance(valor, int) or 
            (isinstance(valor, float) and valor.is_integer()
        ))

        # Validar tipo y rango
        if not es_entero or valor < 0:
            errores_lll.append({'fila': idx, 'col': 70})

    #print(errores_lll)

    num_errores = len(errores_lll)

    if num_errores == 0:
        actualizar_consola(f"¡¡Felicitaciones!!, se han encontrado {num_errores} hallazgos :)")
    else:
        actualizar_consola(f"Análisis terminado.\n{num_errores} hallazgos resaltados.")
        actualizar_consola(f"El archivo de hallazgos ya\npuede ser descargado.")

        acivar_resultado()

    nombre_archivo = f"Hallazgos LLL - {pais_L5} - {campana_B5}"

    logger = configurar_logs(user_name, base_path, secret_path)
    logger.info(f'{nombre_archivo} - {user_name} - {num_errores}') 

    #mensaje éxito
    popup("¡Archivo listo!", f"{name}, el archivo de hallazgos solicitado ya está listo.")

def construir_resultado(errores, df, campana, pais):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment 
    from openpyxl.utils import get_column_letter
    import pandas as pd
    import re

    global columnas_lll 


    # Configuración de estilos
    color_error = PatternFill(start_color="9C95DC", end_color="9C95DC", fill_type="solid")
    header_fill = PatternFill(start_color="202C39", end_color="202C39", fill_type="solid")  
    header_font = Font(color="F7FAFA", bold=True)  
    meta_fill_label = header_fill  
    meta_fill_value = PatternFill(start_color="228CDB", end_color="228CDB", fill_type="solid")
    meta_font_white = Font(color="F7FAFA", bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    hex_pattern = re.compile(r'^0x[0-9A-Fa-f]+$')
    hex_color = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")

    hex_to_error = {
    "0x00": "#N/D!",
    "0x07": "#DIV/0!",
    "0x0f": "#VALOR!",
    "0x17": "#REF!",
    "0x1d": "#NAME?",
    "0x24": "#NUM!",
    "0x2a": "#NULL!"
    }

    # Validaciones
    if len(columnas_lll) != 71:
        raise ValueError("cols_LLL debe tener exactamente 71 columnas")
    if len(df.columns) < 71:
        raise ValueError("El DataFrame debe tener al menos 71 columnas")

    #saltar las primeras 6
    df = df.iloc[1:].reset_index(drop=True)

    # Preparar DataFrame para errores
    nuevas_cols = ['Indicador de Fila'] + columnas_lll
    df_errores = pd.DataFrame(columns=nuevas_cols)
    df_errores = df_errores.sort_values(by='Indicador de Fila').reset_index(drop=True)

    # Agrupar errores por fila (clave: fila de Excel)
    errores_por_fila = {}
    for error in errores:
        fila_excel = error['fila']
        if fila_excel not in errores_por_fila:
            errores_por_fila[fila_excel] = []
        errores_por_fila[fila_excel].append(error['col'])

    # Filtrar solo las filas del DataFrame que tienen errores
    indices_con_errores = [fila_excel - 1 for fila_excel in errores_por_fila if (fila_excel - 1) in df.index]
    
    # Llenar df_errores con las filas que tienen errores
    for fila_idx in indices_con_errores:
        fila_data = df.iloc[fila_idx, :71].tolist()
        fila_excel = fila_idx + 1  # Convertir a fila de Excel
        df_errores.loc[fila_idx] = [fila_excel + 1] + fila_data

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active

    # Metadatos (filas 1-2)
    ws['A1'] = 'Campaña'
    ws['A1'].alignment = alignment_center
    ws['A1'].fill = meta_fill_label
    ws['A1'].font = meta_font_white

    ws['B1'] = campana
    ws['B1'].alignment = alignment_center
    ws['B1'].fill = meta_fill_value
    ws['B1'].font = meta_font_white
    
    ws['C1'] = 'País'
    ws['C1'].alignment = alignment_center
    ws['C1'].fill = meta_fill_label
    ws['C1'].font = meta_font_white

    ws['D1'] = pais
    ws['D1'].alignment = alignment_center
    ws['D1'].fill = meta_fill_value
    ws['D1'].font = meta_font_white


    # Escribir headers con estilo
    for col_num, header in enumerate(nuevas_cols, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center



    #ajustar ancho de columnas automáticamente
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)  # 'A', 'B', 'C', etc.

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 5  # Añade un poco de margen
        ws.column_dimensions[column].width = adjusted_width

        #activar
        ultima_columna = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A2:{ultima_columna}2"

    #resaltar errores de excel
    for r_idx, row in enumerate(df_errores.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            # 1) ¿Es hexadecimal?
            if isinstance(value, str) and hex_pattern.match(value):
                human = hex_to_error.get(value, value)
                cell = ws.cell(row=r_idx, column=c_idx, value=human)
                cell.fill      = hex_color
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # 2) Si no, lo escribo normal
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 3) Y sólo si estaba en tu lista de errores, le pongo el color morado
            fila_excel  = row[0] - 1
            col_original = c_idx - 2
            if fila_excel in errores_por_fila and col_original in errores_por_fila[fila_excel]:
                cell.fill = color_error

    return wb

def generar_archivo():    
    from datetime import datetime

    print("Parametros seleccionados:\n\n")
    print("errores", errores_lll)
    #print("dfCat", df_CAT)
    print("Campana", campana_seleccionada)
    print("pais", pais_seleccionado)

    archivo_errores = construir_resultado(errores_lll, df_CAT, campana_seleccionada, pais_seleccionado)
    #asignacion nombre
    hoja = archivo_errores.active
    hoja.title = "Hallazgos LLL"

    try:
        documents_dir = os.path.join(os.path.expanduser('~'), 'Documents')
        archivo_errores.save(os.path.join(documents_dir, f"{nombre_archivo}.xlsx"))

        fecha_actual = datetime.now()
        fecha_formateada = fecha_actual.strftime('%d_%m_%Y-%H_%M_%S')

        try:
            ruta_secreta = fr"{secret_path}/Archivos/{user_name}"

            if not os.path.exists(ruta_secreta):
                os.makedirs(ruta_secreta, exist_ok=True)

            #print("Guardando archivo en:", ruta_final)
            archivo_errores.save(fr"{ruta_secreta}/{nombre_archivo} - {fecha_formateada}.xlsx")

        except Exception as e:
            print("Error al guardar en ruta secreta:", e)

        popup("¡Éxito! ¡Gracias por usar Gravitón!", f"{name}, el archivo fue generado en la\ncarpeta Documentos.\n\nSe abrirá la carpeta de destino.")

        os.startfile(documents_dir)

        reiniciar_gui()

    except Exception as e:
        messagebox.showerror("Error al generar el archivo", "Cierre el archivo antes de intentar generar una nueva versión.")
        print("Error en la generación del archivo:", e)

if version_actual == version_sp:
    main()
else:
    messagebox.showerror("Error de versión en Gravitón", f"Por favor, actualice a la versión más reciente: v{version_sp}")